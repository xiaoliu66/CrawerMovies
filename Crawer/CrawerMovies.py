import time
from openpyxl import Workbook
import requests
from bs4 import BeautifulSoup
import os
import lxml
import pyautogui

url = 'https://www.dytt8.net/index.htm'
context = requests.get(url)
# 解决乱码 原网站用的是gb2312，用了之后个别字体仍然是乱码。换成了GBK 兼容GB 2312 编码,为GB 2312 的升级版本。
context.encoding = 'gbk'
html = context.text
# 获取所有的a标签
target = BeautifulSoup(html, 'lxml').find('div', class_='co_content2').find_all('a')

base_url = 'https://www.dytt8.net'
beginTime = time.time()
print("任务开始--->" + time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))
num = 0

# 实例化
wb = Workbook()
# 激活 worksheet
ws = wb.active

ws['A1'] = "日期"
ws['B1'] = "电影名称"
ws['C1'] = "网址"
ws['D1'] = "磁力链接"
for a in target:
    tempLink = a.get('href')
    # 过滤
    if tempLink.find("20160320") >= 0:
        continue
    # 截取链接中的日期
    n = tempLink.find("2")
    date = tempLink[n:n + 8]

    href = base_url + tempLink
    context1 = requests.get(href)
    context1.encoding = 'gbk'
    detail_html = context1.text
    # 获取电影标题
    detail_title = BeautifulSoup(detail_html, 'lxml').find('div', class_='title_all').find('h1')

    # 获取电影的种子链接
    target = BeautifulSoup(detail_html, 'lxml').find('div', id='Zoom').find('a')
    magnet = target['href']

    ws.cell(row=num + 2, column=1).value = date
    ws.cell(row=num + 2, column=2).value = detail_title.text
    ws.cell(row=num + 2, column=3).value = href
    ws.cell(row=num + 2, column=4).value = magnet

    print(href + " ---> " + detail_title.text + "--->" + magnet)

    num = num + 1

endTime = time.time()
useTime = (endTime - beginTime)
print(useTime)
print(num)
print("任务结束--->" + time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))

# 将结果保存到excel中
wb.save('movies.xlsx')
