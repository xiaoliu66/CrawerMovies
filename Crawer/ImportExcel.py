import time
import uuid
import pymysql
from openpyxl import Workbook
from openpyxl import load_workbook

wb = load_workbook('./movies.xlsx')
wb.guess_types = True  # 猜测格式类型
ws = wb.active

# 打开数据库连接
db = pymysql.connect(host="localhost", port=3306, user="root", password="123456", db="test")

row_number = 2  # 开始读取的行号
while (ws['a' + str(row_number)].value is not None):
    # 使用 cursor() 方法创建一个游标对象 cursor
    cursor = db.cursor()

    uid = str(uuid.uuid4()).replace('-', '')
    date = ws['a' + str(row_number)].value
    name = ws['b' + str(row_number)].value
    url = ws['c' + str(row_number)].value
    magnet = ws['d' + str(row_number)].value
    create_time = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
    modify_time = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
    is_delete = 'N'

    # SQL 插入语句
    # sql = "INSERT INTO t_web_crawler(id,date, name, url, magnet,context,create_time,modify_time,is_delete)/
    #         VALUES('%s', '%s', % s, '%s', % s)" %
    sql = "INSERT INTO t_web_crawler(id, date, name, url, magnet, create_time, modify_time, is_delete)\
           VALUES ('%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s')" % \
    (uid,date,name,url,magnet,create_time,modify_time,is_delete)
    # print(sql)
    try:
        # 执行sql语句
        cursor.execute(sql)
        # 提交到数据库执行
        db.commit()
        print(uid + "：插入成功！")
    except:
        # 如果发生错误则回滚
        print("插入出错，回滚事务！")
        db.rollback()
    row_number += 1
# 关闭数据库连接
db.close()
